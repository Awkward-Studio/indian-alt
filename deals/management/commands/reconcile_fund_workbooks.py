from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from accounts.models import Profile
from contacts.models import Contact
from deals.models import Deal, FundClassificationSourceType, FundClassificationState
from deals.services.receipt_date_evidence import ReceiptDateEvidenceService


DEFAULT_WORKBOOKS = (
    ("1. Fund I.xlsx", "FUND1"),
    ("2. Fund II.xlsx", "FUND2"),
    ("3. Fund III.xlsx", "FUND3"),
)
MISSING_VALUES = {"", "-", "nan", "none", "n/a", "na", "not specified"}
EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)
INTERNAL_EMAIL_DOMAINS = {"india-alt.com"}
WORKBOOK_FIELDS = {
    "Deal Status", "Date of Receipt", "Deal Name", "Source",
    "Funding Ask (INR MILLION)", "Deal Team", "Industry", "Sector",
    "Is Female Led", "Management Meeting", "Business Proposal Stage",
    "IC Stage", "Next Steps", "City", "Contacts", "Reasons for Passing",
    "Summary", "Details", "Company Info", "Fund",
}
TEXT_FIELD_MAP = {
    "source": "legacy_investment_bank",
    "industry": "industry",
    "sector": "sector",
    "city": "city",
    "summary": "deal_summary",
    "details": "deal_details",
    "company_info": "company_details",
    "next_steps": "comments",
}
BOOLEAN_FIELD_MAP = {
    "is_female_led": "is_female_led",
    "management_meeting": "management_meeting",
    "business_proposal_stage": "business_proposal_stage",
    "ic_stage": "ic_stage",
}


def clean(value) -> str:
    text = str(value or "").strip()
    return "" if text.lower() in MISSING_VALUES else text


def clean_funding_ask(value) -> str:
    text = clean(value)
    return "" if text.lower() in {"0", "0.0"} else text


def parse_bool(value) -> bool | None:
    if isinstance(value, bool):
        return value
    text = clean(value).casefold()
    if text in {"true", "yes", "1"}:
        return True
    if text in {"false", "no", "0"}:
        return False
    return None


def extract_external_emails(value) -> list[str]:
    emails = []
    for email in EMAIL_RE.findall(clean(value)):
        normalized = email.casefold()
        if normalized.rsplit("@", 1)[-1] in INTERNAL_EMAIL_DOMAINS:
            continue
        if normalized not in emails:
            emails.append(normalized)
    return emails


def split_team_initials(value) -> list[str]:
    return [
        item.strip().upper()
        for item in clean(value).split(",")
        if item.strip()
    ]


def normalized_title(value) -> str:
    text = unicodedata.normalize("NFKD", clean(value)).casefold()
    return re.sub(r"[^a-z0-9]+", "", text)


def parse_receipt_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    if not text:
        return None
    for pattern in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


class Command(BaseCommand):
    help = (
        "Reconcile Fund I/II/III workbook receipt dates, pass reasons, and funding asks. "
        "Dry-run by default; exact unique fund/title matches only."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--source-dir",
            default=".",
            help="Directory containing the selected fund workbooks.",
        )
        parser.add_argument(
            "--fund",
            action="append",
            choices=["FUND1", "FUND2", "FUND3"],
            help="Fund to reconcile. Repeat for multiple funds; defaults to all three.",
        )
        parser.add_argument("--apply", action="store_true", help="Persist safe backfills.")

    def handle(self, *args, **options):
        source_dir = Path(options["source_dir"]).resolve()
        selected_funds = set(options["fund"] or ["FUND1", "FUND2", "FUND3"])
        workbook_specs = [
            (source_dir / name, fund)
            for name, fund in DEFAULT_WORKBOOKS
            if fund in selected_funds
        ]
        missing = [str(path) for path, _ in workbook_specs if not path.is_file()]
        if missing:
            raise CommandError(f"Workbook(s) not found: {', '.join(missing)}")

        workbook_rows = defaultdict(list)
        stats = defaultdict(int)
        examples = defaultdict(list)

        for path, expected_fund in workbook_specs:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [clean(value) for value in next(rows)]
            required = WORKBOOK_FIELDS
            missing_headers = required.difference(headers)
            if missing_headers:
                raise CommandError(f"{path.name} is missing columns: {sorted(missing_headers)}")
            positions = {name: headers.index(name) for name in required}

            for row_number, row in enumerate(rows, start=2):
                stats["workbook_rows"] += 1
                fund = clean(row[positions["Fund"]]).upper()
                title = clean(row[positions["Deal Name"]])
                key = (fund, normalized_title(title))
                if fund != expected_fund or not key[1]:
                    stats["skipped_invalid_workbook_rows"] += 1
                    continue
                workbook_rows[key].append(
                    {
                        "workbook": path.name,
                        "row": row_number,
                        "title": title,
                        "received_at": parse_receipt_date(row[positions["Date of Receipt"]]),
                        "reason": clean(row[positions["Reasons for Passing"]]),
                        "funding_ask": clean_funding_ask(
                            row[positions["Funding Ask (INR MILLION)"]]
                        ),
                        "status": clean(row[positions["Deal Status"]]),
                        "source": clean(row[positions["Source"]]),
                        "deal_team": split_team_initials(row[positions["Deal Team"]]),
                        "industry": clean(row[positions["Industry"]]),
                        "sector": clean(row[positions["Sector"]]),
                        "is_female_led": parse_bool(row[positions["Is Female Led"]]),
                        "management_meeting": parse_bool(
                            row[positions["Management Meeting"]]
                        ),
                        "business_proposal_stage": parse_bool(
                            row[positions["Business Proposal Stage"]]
                        ),
                        "ic_stage": parse_bool(row[positions["IC Stage"]]),
                        "next_steps": clean(row[positions["Next Steps"]]),
                        "city": clean(row[positions["City"]]),
                        "contact_emails": extract_external_emails(
                            row[positions["Contacts"]]
                        ),
                        "summary": clean(row[positions["Summary"]]),
                        "details": clean(row[positions["Details"]]),
                        "company_info": clean(row[positions["Company Info"]]),
                    }
                )

        contact_map = defaultdict(list)
        for contact in Contact.objects.exclude(email__isnull=True).only("id", "email"):
            contact_map[clean(contact.email).casefold()].append(contact)

        profile_map = defaultdict(dict)
        for profile in Profile.objects.only("id", "initials", "name"):
            for candidate in (profile.initials, profile.name):
                key = clean(candidate).upper()
                if key:
                    profile_map[key][profile.id] = profile

        database_deals = defaultdict(list)
        for deal in Deal.objects.filter(fund__in=selected_funds).only(
            "id", "title", "fund", "received_at", "reasons_for_passing",
            "rejection_reason", "deal_status", "current_phase", "funding_ask",
            "legacy_investment_bank", "industry", "sector", "city",
            "deal_summary", "deal_details", "company_details", "comments",
            "is_female_led", "management_meeting", "business_proposal_stage",
            "ic_stage", "primary_contact",
            "fund_classification_state", "fund_classification_source_type",
            "fund_classification_source_id",
        ):
            database_deals[(clean(deal.fund).upper(), normalized_title(deal.title))].append(deal)
            stats["database_deals"] += 1

        pending = []
        all_keys = set(workbook_rows) | set(database_deals)
        for key in all_keys:
            source_rows = workbook_rows.get(key, [])
            deals = database_deals.get(key, [])
            if len(source_rows) != 1 or len(deals) != 1:
                if not source_rows:
                    stats["database_without_workbook_match"] += len(deals)
                elif not deals:
                    stats["workbook_without_database_match"] += len(source_rows)
                else:
                    stats["ambiguous_keys"] += 1
                    if len(examples["ambiguous"]) < 10:
                        examples["ambiguous"].append({
                            "fund": key[0],
                            "titles": [item["title"] for item in source_rows],
                            "database_ids": [str(deal.id) for deal in deals],
                        })
                continue

            source = source_rows[0]
            deal = deals[0]
            stats["unique_matches"] += 1
            changes = {}
            contact_ids_to_add = set()
            profile_ids_to_add = set()
            receipt_date_suggestion = None

            workbook_source_id = f'{source["workbook"]}:row:{source["row"]}'
            if (
                deal.fund_classification_state != FundClassificationState.EXPLICIT
                or deal.fund_classification_source_type != FundClassificationSourceType.WORKBOOK
                or deal.fund_classification_source_id != workbook_source_id
            ):
                changes.update({
                    'fund_classification_state': FundClassificationState.EXPLICIT,
                    'fund_classification_source_type': FundClassificationSourceType.WORKBOOK,
                    'fund_classification_source_id': workbook_source_id,
                })
                stats['fund_provenance_to_backfill'] += 1

            source_date = source["received_at"]
            if source_date and deal.received_at is None:
                receipt_date_suggestion = {
                    'proposed_date': source_date,
                    'source_type': 'WORKBOOK',
                    'source_id': workbook_source_id,
                    'evidence': {
                        'workbook': source['workbook'],
                        'row': source['row'],
                        'fund': deal.fund,
                        'deal_title': source['title'],
                        'date_of_receipt': source_date.isoformat(),
                    },
                    'confidence': 1.0,
                }
                stats["date_suggestions_to_create"] += 1
            elif source_date and deal.received_at != source_date:
                stats["date_conflicts_preserved"] += 1
                if len(examples["date_conflicts"]) < 10:
                    examples["date_conflicts"].append({
                        "deal_id": str(deal.id),
                        "title": deal.title,
                        "database": deal.received_at.isoformat(),
                        "workbook": source_date.isoformat(),
                    })
            elif not source_date:
                stats["missing_workbook_dates"] += 1

            is_passed = "Passed" in {deal.deal_status, deal.current_phase}
            existing_reason = clean(deal.reasons_for_passing) or clean(deal.rejection_reason)
            if is_passed and source["reason"] and not existing_reason:
                changes["reasons_for_passing"] = source["reason"]
                changes["rejection_reason"] = source["reason"]
                stats["reasons_to_backfill"] += 1
            elif is_passed and not source["reason"] and not existing_reason:
                stats["passed_without_any_reason"] += 1

            source_ask = source["funding_ask"]
            existing_ask = clean_funding_ask(deal.funding_ask)
            if source_ask and not existing_ask:
                changes["funding_ask"] = source_ask
                stats["funding_asks_to_backfill"] += 1
            elif source_ask and existing_ask != source_ask:
                stats["funding_ask_conflicts_preserved"] += 1
                if len(examples["funding_ask_conflicts"]) < 10:
                    examples["funding_ask_conflicts"].append({
                        "deal_id": str(deal.id),
                        "title": deal.title,
                        "database": deal.funding_ask,
                        "workbook": source_ask,
                    })
            elif not source_ask and not existing_ask:
                stats["missing_funding_ask_in_both"] += 1

            for source_key, model_field in TEXT_FIELD_MAP.items():
                source_value = source[source_key]
                existing_value = clean(getattr(deal, model_field))
                if source_value and not existing_value:
                    changes[model_field] = (
                        f"Next steps: {source_value}"
                        if model_field == "comments"
                        else source_value
                    )
                    stats[f"{model_field}_to_backfill"] += 1
                elif source_value and existing_value:
                    normalized_source = clean(source_value)
                    normalized_existing = clean(existing_value)
                    if (
                        model_field == "comments"
                        and normalized_existing == f"Next steps: {normalized_source}"
                    ):
                        continue
                    if normalized_existing != normalized_source:
                        stats[f"{model_field}_conflicts_preserved"] += 1

            for source_key, model_field in BOOLEAN_FIELD_MAP.items():
                source_value = source[source_key]
                existing_value = getattr(deal, model_field)
                if source_value is True and existing_value is False:
                    changes[model_field] = True
                    stats[f"{model_field}_true_to_backfill"] += 1
                elif source_value is False and existing_value is True:
                    stats[f"{model_field}_conflicts_preserved"] += 1

            existing_contact_ids = set(
                deal.additional_contacts.values_list("id", flat=True)
            )
            if deal.primary_contact_id:
                existing_contact_ids.add(deal.primary_contact_id)
            for email in source["contact_emails"]:
                matches = contact_map[email]
                if len(matches) == 1:
                    contact_id = matches[0].id
                    if contact_id not in existing_contact_ids:
                        contact_ids_to_add.add(contact_id)
                        stats["contact_links_to_backfill"] += 1
                elif not matches:
                    stats["contact_emails_unmatched"] += 1
                else:
                    stats["contact_emails_ambiguous"] += 1

            existing_profile_ids = set(
                deal.responsibility.values_list("id", flat=True)
            )
            for initials in source["deal_team"]:
                matches = list(profile_map[initials].values())
                if len(matches) == 1:
                    profile_id = matches[0].id
                    if profile_id not in existing_profile_ids:
                        profile_ids_to_add.add(profile_id)
                        stats["deal_team_links_to_backfill"] += 1
                elif not matches:
                    stats["deal_team_initials_unmatched"] += 1
                else:
                    stats["deal_team_initials_ambiguous"] += 1

            source_status = source["status"].casefold()
            expected_status = (
                "Passed" if source_status == "passed"
                else "1: Deal Sourced" if source_status == "new"
                else ""
            )
            if expected_status and expected_status not in {
                deal.deal_status, deal.current_phase
            }:
                stats["status_conflicts_preserved"] += 1
            elif not expected_status and source["status"]:
                stats["unmapped_workbook_statuses"] += 1

            if changes or contact_ids_to_add or profile_ids_to_add or receipt_date_suggestion:
                pending.append(
                    (deal, changes, contact_ids_to_add, profile_ids_to_add, receipt_date_suggestion)
                )

        if options["apply"]:
            with transaction.atomic():
                for deal, changes, contact_ids, profile_ids, receipt_suggestion in pending:
                    for field, value in changes.items():
                        setattr(deal, field, value)
                    if changes:
                        deal.save(update_fields=list(changes))
                    if contact_ids:
                        deal.additional_contacts.add(*contact_ids)
                    if profile_ids:
                        deal.responsibility.add(*profile_ids)
                    if receipt_suggestion:
                        _suggestion, created = ReceiptDateEvidenceService.propose(
                            deal=deal,
                            **receipt_suggestion,
                        )
                        stats['date_suggestions_created'] += int(created)
                    stats["deals_updated"] += 1

        report = {
            "mode": "apply" if options["apply"] else "dry-run",
            "source_dir": str(source_dir),
            "funds": sorted(selected_funds),
            "stats": dict(sorted(stats.items())),
            "examples": dict(examples),
        }
        self.stdout.write(json.dumps(report, indent=2, default=str))
