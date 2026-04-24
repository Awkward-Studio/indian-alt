from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from django.core.management.base import BaseCommand, CommandError

from contacts.models import Contact
from deals.models import Deal
from deals.services.contact_linking import sync_deal_contact_links


DEFAULT_SOURCE_DIR = Path("data/legacy_dms_files")
DEFAULT_WORKBOOKS = ["3. Fund III.xlsx"]
EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)
SKIP_EMAIL_DOMAINS = {"india-alt.com"}


def clean_value(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "-"}:
        return ""
    return text


def extract_emails(text: str) -> list[str]:
    if not text:
        return []
    seen: list[str] = []
    for email in EMAIL_RE.findall(text):
        normalized = email.strip().lower()
        domain = normalized.rsplit("@", 1)[-1] if "@" in normalized else ""
        if domain in SKIP_EMAIL_DOMAINS:
            continue
        if normalized not in seen:
            seen.append(normalized)
    return seen


class Command(BaseCommand):
    help = "Second-pass contact backfill for legacy workbook deals using email matches from the Contacts column."

    def add_arguments(self, parser):
        parser.add_argument(
            "--source-dir",
            default=str(DEFAULT_SOURCE_DIR),
            help="Directory containing the legacy workbook exports.",
        )
        parser.add_argument(
            "--workbook",
            action="append",
            help="Workbook filename to process. Defaults to the Fund III workbook only.",
        )
        parser.add_argument(
            "--fund",
            default="FUND3",
            help="Only process rows for this fund value. Use empty string to disable filtering.",
        )
        parser.add_argument("--apply", action="store_true", help="Write links to the database.")
        parser.add_argument(
            "--interactive",
            action="store_true",
            help="Prompt for contact selection on each matched row instead of linking automatically.",
        )
        parser.add_argument(
            "--progress-interval",
            type=int,
            default=100,
            help="Emit a progress line every N matched rows.",
        )

    def handle(self, *args, **options):
        source_dir = Path(options["source_dir"])
        workbook_names = options["workbook"] or DEFAULT_WORKBOOKS
        apply = options["apply"]
        interactive = options["interactive"]
        fund_filter = clean_value(options["fund"])
        progress_interval = options["progress_interval"]

        if interactive and not sys.stdin.isatty():
            raise CommandError("--interactive requires a TTY.")

        if not source_dir.exists():
            raise CommandError(f"Source directory not found: {source_dir}")

        workbooks = [source_dir / workbook_name for workbook_name in workbook_names]
        missing = [str(path) for path in workbooks if not path.exists()]
        if missing:
            raise CommandError(f"Workbook(s) not found: {', '.join(missing)}")

        mode = "APPLY" if apply else "DRY-RUN"
        self.stdout.write(f"[{mode}] source_dir={source_dir} workbooks={len(workbooks)} fund_filter={fund_filter or 'ALL'}")

        stats = {
            "rows_seen": 0,
            "rows_matched": 0,
            "rows_skipped_no_email": 0,
            "rows_skipped_internal_only": 0,
            "rows_skipped_no_deal": 0,
            "contacts_matched": 0,
            "contacts_missing": 0,
            "contacts_duplicated": 0,
            "deals_primary_set": 0,
            "deals_updated": 0,
        }

        for workbook_path in workbooks:
            self.stdout.write("")
            self.stdout.write(f"=== {workbook_path.name} ===")
            wb = load_workbook(workbook_path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            headers = [clean_value(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            header_map = {header: index for index, header in enumerate(headers)}

            required_columns = {"Deal Name", "Contacts"}
            missing_columns = [column for column in required_columns if column not in header_map]
            if missing_columns:
                wb.close()
                raise CommandError(f"{workbook_path.name} is missing required columns: {', '.join(missing_columns)}")

            title_idx = header_map["Deal Name"]
            contacts_idx = header_map["Contacts"]
            fund_idx = header_map.get("Fund")

            for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                stats["rows_seen"] += 1
                title = clean_value(row[title_idx] if title_idx < len(row) else "")
                if not title:
                    continue

                row_fund = clean_value(row[fund_idx] if fund_idx is not None and fund_idx < len(row) else "")
                if fund_filter and row_fund and row_fund.upper() != fund_filter.upper():
                    continue

                contacts_raw = clean_value(row[contacts_idx] if contacts_idx < len(row) else "")
                self.stdout.write(
                    f"[ROW] workbook={workbook_path.name} row={row_number} title={title} fund={row_fund or 'N/A'}"
                )
                self.stdout.write(f"      contacts_raw={contacts_raw or 'N/A'}")
                emails = extract_emails(contacts_raw)
                if not emails:
                    if contacts_raw and "@" in contacts_raw:
                        stats["rows_skipped_internal_only"] += 1
                        self.stdout.write(
                            f"[SKIP-INTERNAL-ONLY] workbook={workbook_path.name} row={row_number} "
                            f"title={title} no external banker emails found"
                        )
                    else:
                        stats["rows_skipped_no_email"] += 1
                        self.stdout.write(
                            f"[SKIP-NO-EMAIL] workbook={workbook_path.name} row={row_number} "
                            f"title={title} fund={row_fund or 'N/A'}"
                        )
                    continue
                self.stdout.write(f"      external_emails={emails}")

                deal = self._find_deal(title, row_fund or fund_filter)
                if not deal:
                    stats["rows_skipped_no_deal"] += 1
                    self.stdout.write(
                        f"[SKIP-NO-DEAL] workbook={workbook_path.name} row={row_number} title={title} "
                        f"fund={row_fund or fund_filter or 'N/A'} emails={emails}"
                    )
                    continue

                self.stdout.write(
                    f"[DEAL] matched_deal={deal.title} | deal_id={deal.id} | "
                    f"existing_primary={deal.primary_contact.name if deal.primary_contact else 'none'} | "
                    f"existing_additional={deal.additional_contacts.count()}"
                )

                matched_contacts: list[Contact] = []
                for email in emails:
                    matches = list(Contact.objects.filter(email__iexact=email).order_by("name", "id"))
                    if not matches:
                        stats["contacts_missing"] += 1
                        self.stdout.write(
                            f"[MISS] workbook={workbook_path.name} row={row_number} title={title} "
                            f"deal={deal.title} email={email}"
                        )
                        continue
                    if len(matches) > 1:
                        stats["contacts_duplicated"] += len(matches) - 1
                        self.stdout.write(
                            f"[DUPLICATE] workbook={workbook_path.name} row={row_number} title={title} "
                            f"deal={deal.title} email={email} matched={len(matches)} "
                            f"using={matches[0].name or matches[0].id}"
                        )
                    contact = matches[0]
                    if all(str(existing.id) != str(contact.id) for existing in matched_contacts):
                        matched_contacts.append(contact)
                        stats["contacts_matched"] += 1

                if not matched_contacts:
                    self.stdout.write(
                        f"[SKIP-NO-CONTACT] workbook={workbook_path.name} row={row_number} title={title} "
                        f"deal={deal.title} emails={emails}"
                    )
                    continue

                selected_contacts = matched_contacts
                primary_contact = None
                should_set_primary = False

                if interactive:
                    selected_contacts, primary_contact = self._prompt_for_contacts(
                        workbook_path=workbook_path,
                        row_number=row_number,
                        deal=deal,
                        matched_contacts=matched_contacts,
                        emails=emails,
                    )
                    if not selected_contacts:
                        self.stdout.write(
                            f"[SKIP] workbook={workbook_path.name} row={row_number} title={title} deal={deal.title}"
                        )
                        continue
                else:
                    if deal.primary_contact_id is None:
                        primary_contact = selected_contacts[0]
                        should_set_primary = True
                    else:
                        should_set_primary = False
                        primary_contact = None

                stats["rows_matched"] += 1
                self.stdout.write(
                    f"[MATCH] workbook={workbook_path.name} row={row_number} title={title} "
                    f"deal={deal.title} | fund={deal.fund or 'N/A'} | "
                    f"selected_contacts={[f'{c.name}<{c.email}>' for c in selected_contacts]} | "
                    f"primary={primary_contact.name if primary_contact else 'none'}"
                )

                if apply:
                    self.stdout.write(
                        f"[APPLY] workbook={workbook_path.name} row={row_number} title={title} "
                        f"deal={deal.title} linking {len(selected_contacts)} contact(s)"
                    )
                    sync_deal_contact_links(
                        deal,
                        primary_contact=primary_contact,
                        primary_contact_provided=should_set_primary,
                        additional_contacts=selected_contacts,
                        additional_contacts_provided=True,
                    )
                    if should_set_primary:
                        stats["deals_primary_set"] += 1
                    stats["deals_updated"] += 1

                if progress_interval > 0 and stats["rows_matched"] % progress_interval == 0:
                    self.stdout.write(
                        f"[PROGRESS] rows_seen={stats['rows_seen']} matched={stats['rows_matched']} "
                        f"contacts_matched={stats['contacts_matched']} missing={stats['contacts_missing']}"
                    )

            wb.close()

        self.stdout.write("-" * 88)
        self.stdout.write(
            "Complete. "
            f"rows_seen={stats['rows_seen']} rows_matched={stats['rows_matched']} "
            f"rows_skipped_no_email={stats['rows_skipped_no_email']} rows_skipped_no_deal={stats['rows_skipped_no_deal']} "
            f"rows_skipped_internal_only={stats['rows_skipped_internal_only']} "
            f"contacts_matched={stats['contacts_matched']} contacts_missing={stats['contacts_missing']} "
            f"contacts_duplicated={stats['contacts_duplicated']} deals_primary_set={stats['deals_primary_set']} "
            f"deals_updated={stats['deals_updated']}"
        )

    def _find_deal(self, title: str, fund: str | None) -> Deal | None:
        queryset = Deal.objects.filter(title__iexact=title)
        if fund:
            match = queryset.filter(fund__iexact=fund).first()
            if match:
                return match
        return queryset.order_by("created_at", "id").first()

    def _prompt_for_contacts(
        self,
        *,
        workbook_path: Path,
        row_number: int,
        deal: Deal,
        matched_contacts: list[Contact],
        emails: list[str],
    ) -> tuple[list[Contact], Contact | None]:
        self.stdout.write(
            f"[INTERACTIVE] workbook={workbook_path.name} row={row_number} deal={deal.title} emails={emails}"
        )
        self.stdout.write("Possible contacts:")
        for index, contact in enumerate(matched_contacts, start=1):
            bank_name = contact.bank.name if contact.bank else "N/A"
            self.stdout.write(
                f"  {index}. {contact.name or contact.id} | email={contact.email or 'N/A'} | bank={bank_name}"
            )

        while True:
            raw_choice = input("Select contacts to link (Enter=all, s=skip, 1,2,...): ").strip().lower()
            if raw_choice in {"s", "skip"}:
                return [], None
            if not raw_choice:
                selected = list(matched_contacts)
                break
            indexes = [part.strip() for part in raw_choice.split(",") if part.strip()]
            if not all(part.isdigit() for part in indexes):
                self.stdout.write("Enter comma-separated numbers, Enter for all, or s to skip.")
                continue
            selected_indexes = sorted({int(part) for part in indexes})
            if not selected_indexes or any(idx < 1 or idx > len(matched_contacts) for idx in selected_indexes):
                self.stdout.write("One or more selected numbers are out of range.")
                continue
            selected = [matched_contacts[idx - 1] for idx in selected_indexes]
            break

        if not selected:
            return [], None

        primary_contact = None
        if deal.primary_contact_id is None:
            if len(selected) == 1:
                primary_contact = selected[0]
            else:
                primary_contact = selected[0]
                self.stdout.write(
                    f"[PRIMARY-DEFAULT] workbook={workbook_path.name} row={row_number} deal={deal.title} "
                    f"defaulting primary to {primary_contact.name or primary_contact.id}"
                )

        self.stdout.write(
            f"[SELECTED] workbook={workbook_path.name} row={row_number} deal={deal.title} "
            f"contacts={[f'{c.name}<{c.email}>' for c in selected]}"
            + (f" | primary={primary_contact.name if primary_contact else 'none'}" if selected else "")
        )
        return selected, primary_contact
