from collections import defaultdict
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from contacts.models import Contact
from deals.management.commands.reconcile_fund_workbooks import (
    DEFAULT_WORKBOOKS,
    clean,
    extract_external_emails,
    normalized_title,
)
from deals.models import Deal, DealFieldProvenance
from deals.services.field_provenance import serializable_field_value


class Command(BaseCommand):
    help = "Synchronize every unambiguous external Fund-workbook contact to matching deal rows."

    def add_arguments(self, parser):
        parser.add_argument("--source-dir", default="data/legacy_dms_files")
        parser.add_argument("--apply", action="store_true")

    @transaction.atomic
    def handle(self, *args, **options):
        source_dir = Path(options["source_dir"]).resolve()
        rows = defaultdict(list)
        for filename, expected_fund in DEFAULT_WORKBOOKS:
            path = source_dir / filename
            if not path.is_file():
                raise CommandError(f"Workbook not found: {path}")
            sheet = load_workbook(path, read_only=True, data_only=True).active
            iterator = sheet.iter_rows(values_only=True)
            headers = [clean(value) for value in next(iterator)]
            positions = {header: index for index, header in enumerate(headers)}
            for row_number, row in enumerate(iterator, start=2):
                fund = clean(row[positions["Fund"]]).upper()
                title = clean(row[positions["Deal Name"]])
                emails = extract_external_emails(row[positions["Contacts"]])
                if fund == expected_fund and title and emails:
                    rows[(fund, normalized_title(title))].append({
                        "emails": tuple(emails),
                        "source_id": f"{filename}:row:{row_number}",
                    })

        deals = defaultdict(list)
        for deal in Deal.objects.filter(fund__in=["FUND1", "FUND2", "FUND3"]).select_related("bank", "primary_contact"):
            deals[(clean(deal.fund).upper(), normalized_title(deal.title))].append(deal)

        contacts_created = deals_updated = provenance_created = 0
        unmatched_rows = conflicting_rows = 0
        for key, source_rows in rows.items():
            target_deals = deals.get(key, [])
            if not target_deals:
                unmatched_rows += len(source_rows)
                continue
            email_sets = {item["emails"] for item in source_rows}
            if len(email_sets) != 1:
                conflicting_rows += len(source_rows)
                continue
            emails = list(next(iter(email_sets)))
            source_id = source_rows[0]["source_id"]
            workbook_contacts = []
            for email in emails:
                contact = Contact.objects.filter(email__iexact=email).order_by("created_at", "id").first()
                if contact is None and options["apply"]:
                    local_name = email.split("@", 1)[0].replace(".", " ").replace("_", " ").replace("-", " ")
                    contact = Contact.objects.create(
                        name=" ".join(part.capitalize() for part in local_name.split()),
                        email=email,
                        contact_type=Contact.ContactType.BANKER,
                    )
                    contacts_created += 1
                if contact is not None:
                    workbook_contacts.append(contact)
            if not options["apply"]:
                deals_updated += len(target_deals)
                continue
            if not workbook_contacts:
                continue
            primary = workbook_contacts[0]
            for deal in target_deals:
                previous = deal.primary_contact
                deal.primary_contact = primary
                deal.save(update_fields=["primary_contact"])
                deal.additional_contacts.set(workbook_contacts[1:])
                DealFieldProvenance.objects.create(
                    deal=deal,
                    field_name="primary_contact",
                    source_type=DealFieldProvenance.SourceType.SHEET,
                    source_id=source_id,
                    previous_value=serializable_field_value(previous),
                    value=str(primary.id),
                )
                deals_updated += 1
                provenance_created += 1

        if not options["apply"]:
            transaction.set_rollback(True)
        action = "Updated" if options["apply"] else "Would update"
        self.stdout.write(
            f"{action} {deals_updated} deal contact links; created {contacts_created} contacts and "
            f"{provenance_created} Sheet provenance rows; {unmatched_rows} workbook rows unmatched; "
            f"{conflicting_rows} workbook rows had conflicting contact lists."
        )
