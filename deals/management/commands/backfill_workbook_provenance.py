from collections import Counter, defaultdict
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from deals.management.commands.reconcile_fund_workbooks import (
    DEFAULT_WORKBOOKS,
    clean,
    clean_funding_ask,
    extract_external_emails,
    normalized_title,
    parse_bool,
    parse_receipt_date,
    split_team_initials,
)
from deals.models import Deal, DealFieldProvenance
from deals.services.field_provenance import serializable_field_value
from contacts.models import Contact


def canonical_workbook_status(value):
    status = clean(value).casefold()
    if "passed" in status:
        return "Passed"
    if "portfolio" in status:
        return "Portfolio"
    if "invested" in status:
        return "Invested"
    return "1: Deal Sourced"


class Command(BaseCommand):
    help = "Backfill missing field provenance from exact Fund I/II/III workbook matches."

    def add_arguments(self, parser):
        parser.add_argument("--source-dir", default="data/legacy_dms_files")
        parser.add_argument("--apply", action="store_true")
        parser.add_argument("--show-conflicts", action="store_true")
        parser.add_argument("--overwrite-conflicts", action="store_true")
        parser.add_argument("--sync-contacts", action="store_true")

    @transaction.atomic
    def handle(self, *args, **options):
        source_dir = Path(options["source_dir"]).resolve()
        workbook_rows = defaultdict(list)

        for filename, expected_fund in DEFAULT_WORKBOOKS:
            path = source_dir / filename
            if not path.is_file():
                raise CommandError(f"Workbook not found: {path}")
            sheet = load_workbook(path, read_only=True, data_only=True).active
            rows = sheet.iter_rows(values_only=True)
            headers = [clean(value) for value in next(rows)]
            positions = {header: index for index, header in enumerate(headers)}
            for row_number, row in enumerate(rows, start=2):
                title = clean(row[positions["Deal Name"]])
                fund = clean(row[positions["Fund"]]).upper()
                if fund != expected_fund or not normalized_title(title):
                    continue
                values = {
                    "title": title,
                    "fund": fund,
                    "received_at": parse_receipt_date(row[positions["Date of Receipt"]]),
                    "funding_ask": clean_funding_ask(row[positions["Funding Ask (INR MILLION)"]]),
                    "legacy_investment_bank": clean(row[positions["Source"]]),
                    "bank_name": clean(row[positions["Source"]]),
                    "industry": clean(row[positions["Industry"]]),
                    "sector": clean(row[positions["Sector"]]),
                    "city": clean(row[positions["City"]]),
                    "is_female_led": parse_bool(row[positions["Is Female Led"]]),
                    "current_phase": canonical_workbook_status(row[positions["Deal Status"]]),
                    "deal_status": canonical_workbook_status(row[positions["Deal Status"]]),
                    "_contact_emails": extract_external_emails(row[positions["Contacts"]]),
                    "_deal_team": split_team_initials(row[positions["Deal Team"]]),
                }
                workbook_rows[(fund, normalized_title(title))].append(
                    (f"{filename}:row:{row_number}", values)
                )

        deals = defaultdict(list)
        for deal in Deal.objects.filter(fund__in=["FUND1", "FUND2", "FUND3"]).select_related("bank", "primary_contact").prefetch_related("responsibility"):
            deals[(clean(deal.fund).upper(), normalized_title(deal.title))].append(deal)

        existing = set(
            DealFieldProvenance.objects.values_list("deal_id", "field_name")
        )
        records = []
        matched_deals = ambiguous_matches = value_conflicts = 0
        contacts_created = contact_links_updated = 0
        conflicts_by_field = Counter()
        conflicts_by_fund = Counter()
        conflict_examples = defaultdict(list)
        for key in set(workbook_rows) & set(deals):
            if len(workbook_rows[key]) != 1 or len(deals[key]) != 1:
                ambiguous_matches += 1
                continue
            source_id, values = workbook_rows[key][0]
            deal = deals[key][0]
            matched_deals += 1
            contact_emails = values.pop("_contact_emails", [])
            deal_team = values.pop("_deal_team", [])
            if deal.bank_id and clean(deal.bank.name).casefold() == clean(values.get("legacy_investment_bank")).casefold():
                values["bank"] = deal.bank
            if deal.primary_contact_id and clean(deal.primary_contact.email).casefold() in contact_emails:
                values["primary_contact"] = deal.primary_contact
            assigned_initials = {
                clean(profile.initials).upper()
                for profile in deal.responsibility.all()
                if clean(profile.initials)
            }
            if assigned_initials and assigned_initials == set(deal_team):
                values["responsibility"] = sorted(str(profile.id) for profile in deal.responsibility.all())

            if options["apply"] and options["sync_contacts"] and contact_emails:
                workbook_contacts = []
                for email in contact_emails:
                    contact = Contact.objects.filter(email__iexact=email).order_by("created_at", "id").first()
                    if contact is None:
                        local_name = email.split("@", 1)[0].replace(".", " ").replace("_", " ").replace("-", " ")
                        contact = Contact.objects.create(
                            name=" ".join(part.capitalize() for part in local_name.split()),
                            email=email,
                            bank=deal.bank,
                            contact_type=Contact.ContactType.BANKER,
                        )
                        contacts_created += 1
                    elif contact.contact_type == Contact.ContactType.OTHER:
                        contact.contact_type = Contact.ContactType.BANKER
                        contact.save(update_fields=["contact_type"])
                    workbook_contacts.append(contact)

                previous_primary = deal.primary_contact
                workbook_primary = workbook_contacts[0]
                if deal.primary_contact_id != workbook_primary.id:
                    deal.primary_contact = workbook_primary
                    deal.save(update_fields=["primary_contact"])
                deal.additional_contacts.set(workbook_contacts[1:])
                records.append(DealFieldProvenance(
                    deal=deal,
                    field_name="primary_contact",
                    source_type=DealFieldProvenance.SourceType.SHEET,
                    source_id=source_id,
                    previous_value=serializable_field_value(previous_primary),
                    value=str(workbook_primary.id),
                ))
                contact_links_updated += 1
            for field_name, workbook_value in values.items():
                if workbook_value in (None, ""):
                    continue
                database_value = values[field_name] if field_name == "responsibility" else getattr(deal, field_name)
                comparable_database = clean_funding_ask(database_value) if field_name == "funding_ask" else database_value
                comparable_workbook = clean_funding_ask(workbook_value) if field_name == "funding_ask" else workbook_value
                if comparable_database != comparable_workbook:
                    value_conflicts += 1
                    conflicts_by_field[field_name] += 1
                    conflicts_by_fund[deal.fund] += 1
                    if len(conflict_examples[field_name]) < 3:
                        conflict_examples[field_name].append({
                            "deal": deal.title,
                            "fund": deal.fund,
                            "field": field_name,
                            "database": serializable_field_value(database_value),
                            "workbook": serializable_field_value(workbook_value),
                            "source": source_id,
                        })
                    if not (options["apply"] and options["overwrite_conflicts"]):
                        continue
                    previous_value = database_value
                    setattr(deal, field_name, workbook_value)
                    deal.save(update_fields=[field_name])
                    records.append(DealFieldProvenance(
                        deal=deal,
                        field_name=field_name,
                        source_type=DealFieldProvenance.SourceType.SHEET,
                        source_id=source_id,
                        previous_value=serializable_field_value(previous_value),
                        value=serializable_field_value(workbook_value),
                    ))
                    continue
                if (deal.id, field_name) in existing:
                    continue
                records.append(DealFieldProvenance(
                    deal=deal,
                    field_name=field_name,
                    source_type=DealFieldProvenance.SourceType.SHEET,
                    source_id=source_id,
                    previous_value=None,
                    value=serializable_field_value(database_value),
                ))

        if options["apply"]:
            DealFieldProvenance.objects.bulk_create(records)
        else:
            transaction.set_rollback(True)

        mode = "created" if options["apply"] else "would create"
        conflict_result = (
            f"overwrote {value_conflicts} workbook conflicts"
            if options["apply"] and options["overwrite_conflicts"]
            else f"preserved {value_conflicts} workbook conflicts"
        )
        self.stdout.write(
            f"Matched {matched_deals} deals; {mode} {len(records)} provenance records; "
            f"{conflict_result}; skipped {ambiguous_matches} ambiguous identity matches."
        )
        if options["sync_contacts"]:
            self.stdout.write(f"Created {contacts_created} contacts and updated {contact_links_updated} workbook contact links.")
        if options["show_conflicts"]:
            self.stdout.write(f"Conflicts by field: {dict(conflicts_by_field.most_common())}")
            self.stdout.write(f"Conflicts by fund: {dict(conflicts_by_fund.most_common())}")
            for field_name in conflicts_by_field:
                for example in conflict_examples[field_name]:
                    self.stdout.write(f"CONFLICT {example}")
