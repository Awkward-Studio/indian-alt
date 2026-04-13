#!/usr/bin/env python
import argparse
import json
import os
import sys
from dataclasses import dataclass
from datetime import UTC, date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import django
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
sys.path.insert(0, str(BASE_DIR))
django.setup()

from deals.models import Deal, DealPhase, DealPriority, DealStatus  # noqa: E402


EXPECTED_HEADERS = [
    "Sr No",
    "Deal Status",
    "Date of Receipt",
    "Days Since",
    "Deal Name",
    "Source",
    "Funding Ask (INR MILLION)",
    "Deal Team",
    "Industry",
    "Sector",
    "Is Female Led",
    "Management Meeting",
    "Business Proposal Stage",
    "IC Stage",
    "Next Steps",
    "City",
    "Contacts",
    "Reasons for Passing",
    "Summary",
    "Details",
    "Company Info",
    "Fund",
]

HEADER_MAP = {header.strip().lower(): header for header in EXPECTED_HEADERS}
VALID_STATUSES = {choice for choice, _ in DealStatus.choices}
VALID_PRIORITIES = {choice for choice, _ in DealPriority.choices}


@dataclass
class RowResult:
    row_number: int
    status: str
    deal_name: str
    message: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import deals from an Excel workbook into the Django deal database."
    )
    parser.add_argument("excel_path", help="Path to the .xlsx file to import.")
    parser.add_argument("--sheet", help="Worksheet name to import. Defaults to the active sheet.")
    parser.add_argument("--dry-run", action="store_true", help="Validate and report without writing to the database.")
    parser.add_argument("--report", help="Optional path to write a JSON summary report.")
    return parser.parse_args()


def canonical_header(value: Any) -> str:
    return str(value).strip()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def normalize_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False

    normalized = normalize_text(value).lower()
    if normalized in {"yes", "y", "true", "1"}:
        return True
    if normalized in {"no", "n", "false", "0", ""}:
        return False
    raise ValueError(f"Unsupported boolean value: {value!r}")


def normalize_status(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return DealStatus.NEW

    aliases = {
        "new": DealStatus.NEW,
        "to be passed": DealStatus.TO_BE_PASSED,
        "to be pass": DealStatus.TO_BE_PASS,
        "passed": DealStatus.PASSED,
        "portfolio": DealStatus.PORTFOLIO,
        "invested": DealStatus.INVESTED,
    }
    normalized = aliases.get(text.lower())
    if normalized and normalized in VALID_STATUSES:
        return normalized
    raise ValueError(f"Unsupported deal status: {text}")


def normalize_priority(value: Any = None) -> str:
    text = normalize_text(value)
    if not text:
        return DealPriority.MEDIUM

    aliases = {
        "high": DealPriority.HIGH,
        "medium": DealPriority.MEDIUM,
        "low": DealPriority.LOW,
    }
    normalized = aliases.get(text.lower())
    if normalized and normalized in VALID_PRIORITIES:
        return normalized
    raise ValueError(f"Unsupported priority: {text}")


def normalize_funding_ask(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return "0"

    cleaned = text.replace(",", "")
    try:
        numeric = Decimal(cleaned)
    except InvalidOperation as exc:
        raise ValueError(f"Unsupported funding ask: {value!r}") from exc

    # Source file is expected in INR million, stored value is in INR crores.
    crores = numeric / Decimal("10")
    normalized = crores.quantize(Decimal("0.01"))
    return format(normalized.normalize(), "f")


def normalize_created_at(value: Any) -> datetime | None:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, time.min)
    else:
        text = normalize_text(value)
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                dt = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        else:
            raise ValueError(f"Unsupported receipt date: {value!r}")

    if dt.tzinfo is None:
        return dt.replace(tzinfo=UTC)
    return dt.astimezone(UTC)


def load_rows(excel_path: Path, sheet_name: str | None) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    workbook = load_workbook(excel_path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
      raise ValueError("Workbook is empty.")

    raw_headers = [canonical_header(cell) for cell in rows[0]]
    normalized_headers = [HEADER_MAP.get(header.lower(), header) for header in raw_headers]

    missing = [header for header in EXPECTED_HEADERS if header not in normalized_headers]
    if missing:
        raise ValueError(f"Missing required headers: {', '.join(missing)}")

    indexed_rows: list[tuple[int, dict[str, Any]]] = []
    for row_index, values in enumerate(rows[1:], start=2):
        row_dict = {normalized_headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(normalized_headers))}
        indexed_rows.append((row_index, row_dict))

    return normalized_headers, indexed_rows


def build_payload(row: dict[str, Any]) -> dict[str, Any]:
    deal_name = normalize_text(row["Deal Name"])
    if not deal_name:
        raise ValueError("Deal Name is required.")

    payload = {
        "title": deal_name,
        "deal_status": normalize_status(row.get("Deal Status")),
        "priority": normalize_priority(),
        "current_phase": DealPhase.STAGE_1,
        "legacy_investment_bank": normalize_text(row.get("Source")),
        "funding_ask": normalize_funding_ask(row.get("Funding Ask (INR MILLION)")),
        "industry": normalize_text(row.get("Industry")),
        "sector": normalize_text(row.get("Sector")),
        "is_female_led": normalize_bool(row.get("Is Female Led")),
        "management_meeting": normalize_bool(row.get("Management Meeting")),
        "business_proposal_stage": normalize_bool(row.get("Business Proposal Stage")),
        "ic_stage": normalize_bool(row.get("IC Stage")),
        "comments": normalize_text(row.get("Next Steps")),
        "city": normalize_text(row.get("City")),
        "reasons_for_passing": normalize_text(row.get("Reasons for Passing")),
        "deal_summary": normalize_text(row.get("Summary")),
        "deal_details": normalize_text(row.get("Details")),
        "company_details": normalize_text(row.get("Company Info")),
        "fund": normalize_text(row.get("Fund")) or "FUND3",
        "_created_at": normalize_created_at(row.get("Date of Receipt")),
    }
    return payload


def report_summary(results: list[RowResult]) -> dict[str, Any]:
    summary = {
        "total_rows": len(results),
        "created": sum(1 for result in results if result.status == "created"),
        "skipped": sum(1 for result in results if result.status == "skipped"),
        "failed": sum(1 for result in results if result.status == "failed"),
        "rows": [
            {
                "row_number": result.row_number,
                "status": result.status,
                "deal_name": result.deal_name,
                "message": result.message,
            }
            for result in results
        ],
    }
    return summary


def write_report(report_path: Path, summary: dict[str, Any]) -> None:
    report_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")


def import_rows(rows: list[tuple[int, dict[str, Any]]], dry_run: bool) -> list[RowResult]:
    results: list[RowResult] = []

    for row_number, row in rows:
        deal_name = normalize_text(row.get("Deal Name"))
        try:
            payload = build_payload(row)
            created_at = payload.pop("_created_at", None)

            if Deal.objects.filter(title__iexact=payload["title"]).exists():
                results.append(RowResult(row_number, "skipped", deal_name, "Duplicate deal name."))
                continue

            if dry_run:
                results.append(RowResult(row_number, "created", deal_name, "Validated (dry run)."))
                continue

            deal = Deal.objects.create(**payload)
            if created_at is not None:
                Deal.objects.filter(pk=deal.pk).update(created_at=created_at)
            results.append(RowResult(row_number, "created", deal_name, "Created successfully."))
        except Exception as exc:  # noqa: BLE001
            results.append(RowResult(row_number, "failed", deal_name, str(exc)))

    return results


def print_summary(summary: dict[str, Any]) -> None:
    print("")
    print("Import summary")
    print(f"  Total rows: {summary['total_rows']}")
    print(f"  Created:    {summary['created']}")
    print(f"  Skipped:    {summary['skipped']}")
    print(f"  Failed:     {summary['failed']}")

    if summary["rows"]:
        print("")
        print("Row details")
        for row in summary["rows"]:
            print(f"  Row {row['row_number']}: {row['status'].upper()} - {row['deal_name'] or '(blank deal name)'} - {row['message']}")


def main() -> int:
    args = parse_args()
    excel_path = Path(args.excel_path).expanduser().resolve()

    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}", file=sys.stderr)
        return 1

    try:
        _, rows = load_rows(excel_path, args.sheet)
        results = import_rows(rows, args.dry_run)
        summary = report_summary(results)
        print_summary(summary)
        if args.report:
            write_report(Path(args.report).expanduser().resolve(), summary)
    except Exception as exc:  # noqa: BLE001
        print(f"Import failed: {exc}", file=sys.stderr)
        return 1

    return 0 if summary["failed"] == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
